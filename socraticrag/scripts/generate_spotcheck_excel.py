"""
Generate professor spot-check Excel files for M2 and M3 human validation.

Each workbook has two tabs:
  1. READ ME FIRST  — full instructions, rubric, colour legend
  2. Data sheet     — yellow columns to fill in, blue columns = our automated scores

Outputs:
  data/spotcheck_M2_professor.xlsx
  data/spotcheck_M3_professor.xlsx

Usage:
  python3 scripts/generate_spotcheck_excel.py
"""

import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

M2_INPUT  = Path("data/m2_spot_check.jsonl")
M3_INPUT  = Path("data/m3_spot_check.jsonl")
M2_OUTPUT = Path("data/spotcheck_M2_professor.xlsx")
M3_OUTPUT = Path("data/spotcheck_M3_professor.xlsx")

# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1565C0"   # deep blue  — column headers
C_HEADER_FG  = "FFFFFF"
C_README_HDR = "0D47A1"   # darker blue — README section headers
C_INSTR_BG   = "FFF8E1"   # amber      — banner row
C_INSTR_FG   = "5D4037"
C_SUPPORTED  = "E8F5E9"   # pale green — our verdict
C_NOT_SUP    = "FFEBEE"   # pale red   — our verdict
C_VACUOUS    = "F5F5F5"   # grey
C_YOUR_BG    = "FFF9C4"   # yellow     — professor fills
C_OUR_BG     = "DDEEFF"   # blue       — our automated scores
C_ODD_ROW    = "FAFAFA"
C_EVEN_ROW   = "FFFFFF"
C_STATE = {
    "accurate":      "E8F5E9",
    "erroneous":     "FFEBEE",
    "comprehension": "FFF9C4",
    "confusion":     "F3E5F5",
}
SCORE_COLORS = ["FFCDD2", "FFE0B2", "FFF9C4", "C8E6C9"]  # 0=red … 3=green

# ── Style helpers ─────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def wrap_align(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

_ILLEGAL = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

def safe(value):
    return _ILLEGAL.sub("", str(value)) if value is not None else ""

def set_cell(ws, row, col, value, bg=None, bold=False, fg="000000",
             h="left", v="top", size=10, border=True, italic=False):
    cell = ws.cell(row=row, column=col, value=safe(value))
    if bg:
        cell.fill = fill(bg)
    cell.font      = _font(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = wrap_align(h=h, v=v)
    if border:
        cell.border = thin_border()
    return cell

def score_bg(val):
    try:
        return SCORE_COLORS[int(val)]
    except (TypeError, ValueError, IndexError):
        return "FFFFFF"

# ── LaTeX stripping ───────────────────────────────────────────────────────────

def strip_latex(text):
    if not text:
        return text
    text = re.sub(r'\$\$(.+?)\$\$', lambda m: _latex_expr(m.group(1)), text, flags=re.DOTALL)
    text = re.sub(r'\\\[(.+?)\\\]', lambda m: _latex_expr(m.group(1)), text, flags=re.DOTALL)
    text = re.sub(r'\\\((.+?)\\\)', lambda m: _latex_expr(m.group(1)), text, flags=re.DOTALL)
    text = re.sub(r'\$(.+?)\$',    lambda m: _latex_expr(m.group(1)), text)
    text = re.sub(r'\\[a-zA-Z]+\{([^}]*)\}', r'\1', text)
    text = re.sub(r'\\[a-zA-Z]+', '', text)
    text = re.sub(r'  +', ' ', text)
    return text.strip()

def _latex_expr(expr):
    expr = expr.strip()
    replacements = [
        (r'\\mathcal\{([^}]+)\}', r'\1'), (r'\\mathrm\{([^}]+)\}', r'\1'),
        (r'\\mathbf\{([^}]+)\}',  r'\1'), (r'\\mathbb\{([^}]+)\}',  r'\1'),
        (r'\\text\{([^}]+)\}',    r'\1'), (r'\\hat\{([^}]+)\}',  r'\1_hat'),
        (r'\\tilde\{([^}]+)\}', r'\1_tilde'), (r'\\bar\{([^}]+)\}', r'\1_bar'),
        (r'\\vec\{([^}]+)\}',     r'\1'),
        (r'\\frac\{([^}]+)\}\{([^}]+)\}', r'(\1)/(\2)'),
        (r'\\sqrt\{([^}]+)\}', r'sqrt(\1)'),
        (r'\\sum', 'sum'),    (r'\\prod', 'prod'),  (r'\\int', 'integral'),
        (r'\\infty', 'inf'),  (r'\\alpha', 'alpha'),(r'\\beta', 'beta'),
        (r'\\gamma', 'gamma'),(r'\\delta', 'delta'),(r'\\epsilon', 'epsilon'),
        (r'\\theta', 'theta'),(r'\\lambda', 'lambda'),(r'\\mu', 'mu'),
        (r'\\sigma', 'sigma'),(r'\\pi', 'pi'),      (r'\\tau', 'tau'),
        (r'\\phi', 'phi'),    (r'\\psi', 'psi'),    (r'\\omega', 'omega'),
        (r'\\rightarrow', '->'),(r'\\leftarrow', '<-'),
        (r'\\leq', '<='),    (r'\\geq', '>='),     (r'\\neq', '!='),
        (r'\\approx', '≈'),  (r'\\times', 'x'),    (r'\\cdot', '·'),
        (r'\\ldots', '...'), (r'\\mid', '|'),
        (r'\^\{([^}]+)\}', r'^\1'), (r'_\{([^}]+)\}', r'_\1'),
        (r'\{([^}]*)\}', r'\1'),    (r'\\[a-zA-Z]+', ''), (r'\\', ''),
    ]
    for p, r in replacements:
        expr = re.sub(p, r, expr)
    return expr.strip()

_VERDICT_LABELS = {"ENTAILED": "SUPPORTED", "NOT_ENTAILED": "NOT SUPPORTED"}

def human_verdict(v):
    return _VERDICT_LABELS.get(v, v)

def verdict_color(v):
    return {"SUPPORTED": C_SUPPORTED, "NOT SUPPORTED": C_NOT_SUP}.get(v, C_VACUOUS)

# ── README tab builders ───────────────────────────────────────────────────────

def _readme_section(ws, row, text, is_header=False, is_subheader=False,
                    bg=None, fg="000000", italic=False, indent=0):
    """Write one text block into col B, spanning to col I, return next row."""
    n_cols = 9
    ws.merge_cells(f"B{row}:I{row}")
    prefix = "    " * indent
    cell = ws.cell(row=row, column=2, value=prefix + safe(text))
    cell.font      = _font(bold=is_header or is_subheader,
                           color=(C_README_HDR if is_header else fg),
                           size=(13 if is_header else 11 if is_subheader else 10),
                           italic=italic)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if bg:
        for c in range(1, n_cols + 2):
            ws.cell(row=row, column=c).fill = fill(bg)
    if is_header:
        ws.row_dimensions[row].height = 28
    elif is_subheader:
        ws.row_dimensions[row].height = 20
    else:
        ws.row_dimensions[row].height = 16
    return row + 1


def add_m2_readme(wb):
    ws = wb.create_sheet("READ ME FIRST", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["B"].width = 26

    r = 2
    r = _readme_section(ws, r, "SocraticRAG Benchmark — Expert Validation",
                        is_header=True, bg="E3F2FD")
    r = _readme_section(ws, r, "File M2 — Retrieval Faithfulness  (~20–25 min)",
                        is_subheader=True, bg="E3F2FD")
    r += 1

    r = _readme_section(ws, r, "WHAT THIS IS", is_subheader=True)
    for line in [
        "An AI tutor was given a source text (Context C) and responded to a student question.",
        "My system extracted the factual claims embedded in those responses — called presuppositions —",
        "and checked whether each claim is directly supported by the source text.",
        "Your job: verify whether you agree with those checks.",
    ]:
        r = _readme_section(ws, r, line, indent=1)
    r += 1

    r = _readme_section(ws, r, "HOW TO FILL IN", is_subheader=True)
    for step in [
        "1. Go to the 'M2 Faithfulness' tab.",
        "2. Read the Context C and Tutor Response R for each row.",
        "3. Read the presupposition — a claim extracted from the response.",
        "4. Ask yourself: 'Could a reader derive this claim directly from Context C?'",
        "5. Type SUPPORTED or NOT SUPPORTED in the yellow YOUR_VERDICT column.",
        "6. Skip rows pre-filled with 'N/A — skip' (vacuous responses with no claim).",
    ]:
        r = _readme_section(ws, r, step, indent=1)
    r += 1

    r = _readme_section(ws, r, "VERDICTS", is_subheader=True)
    r = _readme_section(ws, r, "SUPPORTED       — the claim is directly derivable from Context C.",
                        bg=C_SUPPORTED, indent=1)
    r = _readme_section(ws, r, "NOT SUPPORTED   — the claim is absent from or contradicts Context C.",
                        bg=C_NOT_SUP, indent=1)
    r += 1

    r = _readme_section(ws, r, "COLOUR LEGEND", is_subheader=True)
    r = _readme_section(ws, r, "Yellow column  → fill this in (your verdict)", bg=C_YOUR_BG, indent=1)
    r = _readme_section(ws, r, "Grey tint row  → vacuous response, skip", bg=C_VACUOUS, indent=1)
    r += 1

    r = _readme_section(ws, r, "WHAT HAPPENS WITH YOUR SCORES", is_subheader=True)
    for line in [
        "I will compute Cohen's κ between your verdicts and mine to validate the M2 metric for the paper.",
        "No special knowledge of AI is required — the rubric above is self-contained.",
        "If anything is unclear, just reply and I'll clarify.",
    ]:
        r = _readme_section(ws, r, line, indent=1)

    wb.active = ws


def add_m3_readme(wb):
    ws = wb.create_sheet("READ ME FIRST", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["B"].width = 26

    r = 2
    r = _readme_section(ws, r, "SocraticRAG Benchmark — Expert Validation",
                        is_header=True, bg="E3F2FD")
    r = _readme_section(ws, r, "File M3 — Pedagogical Alignment  (~20 min)",
                        is_subheader=True, bg="E3F2FD")
    r += 1

    r = _readme_section(ws, r, "WHAT THIS IS", is_subheader=True)
    for line in [
        "An AI tutor responded to a student who was in one of four cognitive states:",
        "accurate understanding, erroneous belief, comprehension gap, or confusion.",
        "Your job: rate how well the tutor response handled the situation on three dimensions.",
    ]:
        r = _readme_section(ws, r, line, indent=1)
    r += 1

    r = _readme_section(ws, r, "HOW TO FILL IN", is_subheader=True)
    for step in [
        "1. Go to the 'M3 Pedagogical' tab.",
        "2. Read Context C (source text), Student Utterance, and Tutor Response.",
        "3. Rate the response on three dimensions (0–3 each) in the yellow columns.",
        "4. Enter YOUR_P, YOUR_O, YOUR_E. Write the sum manually in YOUR_TOTAL.",
    ]:
        r = _readme_section(ws, r, step, indent=1)
    r += 1

    r = _readme_section(ws, r, "RUBRIC — score each dimension 0 to 3", is_subheader=True)
    r += 1

    r = _readme_section(ws, r, "Perception — did the tutor correctly read the student's cognitive state?",
                        is_subheader=True, fg="1565C0")
    rubric_p = [
        ("0", "Ignores the student's state entirely"),
        ("1", "Acknowledges it but misreads what type of state it is"),
        ("2", "Correctly identifies the state"),
        ("3", "Correctly identifies AND adapts the strategy accordingly"),
    ]
    for score, desc in rubric_p:
        r = _readme_section(ws, r, f"  {score}  —  {desc}",
                            bg=SCORE_COLORS[int(score)], indent=1)
    r += 1

    r = _readme_section(ws, r, "Orchestration — did the tutor choose the right response strategy?",
                        is_subheader=True, fg="1565C0")
    rubric_o = [
        ("0", "Directly gives the answer or solution"),
        ("1", "Hints, but gives too much away"),
        ("2", "Scaffolds appropriately without revealing the answer"),
        ("3", "Fully withholds the answer, guides only through questions"),
    ]
    for score, desc in rubric_o:
        r = _readme_section(ws, r, f"  {score}  —  {desc}",
                            bg=SCORE_COLORS[int(score)], indent=1)
    r += 1

    r = _readme_section(ws, r, "Elicitation — did the tutor draw the student out at the right level?",
                        is_subheader=True, fg="1565C0")
    rubric_e = [
        ("0", "Closed yes/no or rhetorical question"),
        ("1", "Leads the student toward one specific answer"),
        ("2", "Open question that invites reasoning"),
        ("3", "Probes deeper, extends thinking, raises the cognitive level"),
    ]
    for score, desc in rubric_e:
        r = _readme_section(ws, r, f"  {score}  —  {desc}",
                            bg=SCORE_COLORS[int(score)], indent=1)
    r += 1

    r = _readme_section(ws, r, "COLOUR LEGEND", is_subheader=True)
    r = _readme_section(ws, r, "Yellow columns  → fill these in (your scores)", bg=C_YOUR_BG, indent=1)
    r += 1

    r = _readme_section(ws, r, "WHAT HAPPENS WITH YOUR SCORES", is_subheader=True)
    for line in [
        "I will compute Spearman rho between your scores and mine to validate the M3 metric for the paper.",
        "No special knowledge of AI is required — the rubric above is self-contained.",
        "If anything is unclear, just reply and I'll clarify.",
    ]:
        r = _readme_section(ws, r, line, indent=1)

    wb.active = ws


# ── M2 data sheet ─────────────────────────────────────────────────────────────

def build_m2():
    items = []
    with open(M2_INPUT) as f:
        for line in f:
            items.append(json.loads(line))

    wb = Workbook()
    wb.remove(wb.active)          # remove default empty sheet
    add_m2_readme(wb)             # tab 0: READ ME FIRST

    ws = wb.create_sheet("M2 Faithfulness")   # tab 1: data
    ws.freeze_panes = "A3"

    # Yellow = professor; Blue = ours (placed after yellow so professor scores blind first)
    cols = [
        ("row_id",           6),
        ("cognitive_state", 16),
        ("context_C",       58),
        ("tutor_response_R",50),
        ("presup_#",         7),
        ("presupposition",  50),
        ("YOUR_VERDICT",    22),   # yellow — professor fills
        ("notes",           28),
    ]
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row
    YOUR_COL = 7
    for ci, (h, _) in enumerate(cols, 1):
        bg = C_YOUR_BG if ci == YOUR_COL else C_HEADER_BG
        fg = "333333" if ci == YOUR_COL else C_HEADER_FG
        set_cell(ws, 1, ci, h, bg=bg, bold=True, fg=fg, h="center", v="center")
    ws.row_dimensions[1].height = 22

    # Instruction banner
    n = len(cols)
    ws.merge_cells(f"A2:{get_column_letter(n)}2")
    set_cell(ws, 2, 1,
             "Fill the YELLOW column only. Enter SUPPORTED or NOT SUPPORTED. "
             "Skip rows pre-filled with N/A.",
             bg=C_INSTR_BG, fg=C_INSTR_FG, h="left", v="center", size=9)
    ws.row_dimensions[2].height = 28

    row_num = 3
    csv_row_id = 1
    for item in items:
        presups  = item.get("m2_presuppositions", [])
        verdicts = item.get("m2_verdicts", [])
        state    = item.get("cognitive_state", "")
        chunk    = strip_latex(item.get("chunk_text", ""))
        response = strip_latex(item.get("response", ""))
        state_bg = C_STATE.get(state, "FFFFFF")

        if not presups:
            row_bg = C_ODD_ROW if row_num % 2 else C_EVEN_ROW
            set_cell(ws, row_num, 1, csv_row_id,                       bg=row_bg,   h="center")
            set_cell(ws, row_num, 2, state,                             bg=state_bg)
            set_cell(ws, row_num, 3, chunk,                             bg=row_bg)
            set_cell(ws, row_num, 4, response,                          bg=row_bg)
            set_cell(ws, row_num, 5, "—",                               bg=row_bg,   h="center")
            set_cell(ws, row_num, 6, "(no presuppositions — vacuous)",  bg=row_bg)
            set_cell(ws, row_num, 7, "N/A — skip", bg=C_VACUOUS, h="center")
            set_cell(ws, row_num, 8, "",           bg=row_bg)
            ws.row_dimensions[row_num].height = 60
            row_num += 1
            csv_row_id += 1
            continue

        for p_idx, presup in enumerate(presups):
            row_bg = C_ODD_ROW if row_num % 2 else C_EVEN_ROW
            set_cell(ws, row_num, 1, csv_row_id, bg=row_bg,    h="center")
            set_cell(ws, row_num, 2, state,       bg=state_bg)
            set_cell(ws, row_num, 3, chunk,       bg=row_bg)
            set_cell(ws, row_num, 4, response,    bg=row_bg)
            set_cell(ws, row_num, 5, p_idx + 1,  bg=row_bg,    h="center")
            set_cell(ws, row_num, 6, presup,      bg=row_bg)
            set_cell(ws, row_num, 7, "",          bg=C_YOUR_BG, h="center")
            set_cell(ws, row_num, 8, "",          bg=row_bg)
            ws.row_dimensions[row_num].height = 70
            row_num += 1
            csv_row_id += 1

    wb.save(M2_OUTPUT)
    print(f"Saved M2 Excel → {M2_OUTPUT}  ({csv_row_id - 1} rows)")


# ── M3 data sheet ─────────────────────────────────────────────────────────────

def build_m3():
    items = []
    with open(M3_INPUT) as f:
        for line in f:
            items.append(json.loads(line))

    wb = Workbook()
    wb.remove(wb.active)
    add_m3_readme(wb)

    ws = wb.create_sheet("M3 Pedagogical")
    ws.freeze_panes = "A3"

    cols = [
        ("row_id",           6),
        ("cognitive_state", 16),
        ("target_concept",  24),
        ("context_C",       52),
        ("student_utt",     42),
        ("tutor_response",  42),
        ("YOUR_P",          10),   # yellow
        ("YOUR_O",          10),   # yellow
        ("YOUR_E",          10),   # yellow
        ("YOUR_TOTAL",      11),   # yellow
        ("notes",           28),
    ]
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    YOUR_COLS = {7, 8, 9, 10}

    for ci, (h, _) in enumerate(cols, 1):
        bg = C_YOUR_BG if ci in YOUR_COLS else C_HEADER_BG
        fg = "333333"  if ci in YOUR_COLS else C_HEADER_FG
        set_cell(ws, 1, ci, h, bg=bg, bold=True, fg=fg, h="center", v="center")
    ws.row_dimensions[1].height = 22

    n = len(cols)
    ws.merge_cells(f"A2:{get_column_letter(n)}2")
    set_cell(ws, 2, 1,
             "Fill the YELLOW columns only: YOUR_P, YOUR_O, YOUR_E (each 0–3), "
             "then write the sum in YOUR_TOTAL.",
             bg=C_INSTR_BG, fg=C_INSTR_FG, h="left", v="center", size=9)
    ws.row_dimensions[2].height = 28

    row_num = 3
    for idx, item in enumerate(items):
        profile  = item.get("profile", {})
        state    = item.get("cognitive_state", "")
        concept  = profile.get("target_concept", "")
        chunk    = strip_latex(item.get("chunk_text", ""))
        utt      = strip_latex(item.get("utterance", ""))
        resp     = strip_latex(item.get("response", ""))
        p        = item.get("m3_perception", "")
        o        = item.get("m3_orchestration", "")
        el       = item.get("m3_elicitation", "")
        total    = item.get("m3_total", "")
        state_bg = C_STATE.get(state, "FFFFFF")
        row_bg   = C_ODD_ROW if row_num % 2 else C_EVEN_ROW

        set_cell(ws, row_num,  1, idx + 1, bg=row_bg,          h="center")
        set_cell(ws, row_num,  2, state,    bg=state_bg)
        set_cell(ws, row_num,  3, concept,  bg=row_bg)
        set_cell(ws, row_num,  4, chunk,    bg=row_bg)
        set_cell(ws, row_num,  5, utt,      bg=row_bg)
        set_cell(ws, row_num,  6, resp,     bg=row_bg)
        set_cell(ws, row_num,  7, "", bg=C_YOUR_BG, h="center")
        set_cell(ws, row_num,  8, "", bg=C_YOUR_BG, h="center")
        set_cell(ws, row_num,  9, "", bg=C_YOUR_BG, h="center")
        set_cell(ws, row_num, 10, "", bg=C_YOUR_BG, h="center")
        set_cell(ws, row_num, 11, "", bg=row_bg)
        ws.row_dimensions[row_num].height = 90
        row_num += 1

    wb.save(M3_OUTPUT)
    print(f"Saved M3 Excel → {M3_OUTPUT}  ({len(items)} rows)")


if __name__ == "__main__":
    build_m2()
    build_m3()
    print("\nDone.")
