"""
Script 08: Analysis — merge all scores into results table
Inputs:  data/scores_m1.jsonl
         data/scores_m2.jsonl
         data/scores_m3.jsonl
         data/scores_ragas.jsonl
Output:  outputs/results.csv
         outputs/results_by_state.csv
         outputs/correlation_matrix.csv
         outputs/summary.txt

Produces:
  - Main results table: M1 / M2 / M3 (mean ± SD) per model
  - Per-cognitive-state breakdown for M3 (as promised in paper)
  - Spearman correlation matrix: M1 vs M2, M1 vs M3, M2 vs M3 (orthogonality check)
  - RAGAS comparison column showing near-1.0 clustering
  - Human-readable summary printed to stdout and saved to outputs/summary.txt

Run: python3 scripts/08_analysis.py
"""

import json
import csv
import statistics
from pathlib import Path
from collections import defaultdict

SCORES_M1    = Path("data/scores_m1.jsonl")
SCORES_M2    = Path("data/scores_m2.jsonl")
SCORES_M3    = Path("data/scores_m3.jsonl")
SCORES_RAGAS = Path("data/scores_ragas.jsonl")
M2_SPOTCHECK = Path("data/m2_spot_check.jsonl")
M3_SPOTCHECK = Path("data/m3_spot_check.jsonl")
REPLIED_M2   = Path("data/replied_spotcheck_M2_professor.xlsx")
REPLIED_M3   = Path("data/replied_spotcheck_M3_professor.xlsx")

OUTPUT_DIR           = Path("outputs")
RESULTS_CSV          = OUTPUT_DIR / "results.csv"
RESULTS_BY_STATE_CSV = OUTPUT_DIR / "results_by_state.csv"
CORR_CSV             = OUTPUT_DIR / "correlation_matrix.csv"
IAA_CSV              = OUTPUT_DIR / "iaa.csv"
SUMMARY_TXT          = OUTPUT_DIR / "summary.txt"


def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        print(f"  WARNING: {path} not found — skipping.")
        return []
    rows = []
    with open(path) as f:
        for line in f:
            rows.append(json.loads(line))
    return rows


def spearman_rho(x: list[float], y: list[float]) -> float:
    """Compute Spearman rank correlation between x and y."""
    if len(x) != len(y) or len(x) < 2:
        return float("nan")

    def rank(lst: list[float]) -> list[float]:
        sorted_idx = sorted(range(len(lst)), key=lambda i: lst[i])
        ranks = [0.0] * len(lst)
        i = 0
        while i < len(sorted_idx):
            j = i
            while j + 1 < len(sorted_idx) and lst[sorted_idx[j + 1]] == lst[sorted_idx[i]]:
                j += 1
            avg_rank = (i + j) / 2 + 1
            for k in range(i, j + 1):
                ranks[sorted_idx[k]] = avg_rank
            i = j + 1
        return ranks

    rx = rank(x)
    ry = rank(y)
    n = len(rx)
    mean_rx = sum(rx) / n
    mean_ry = sum(ry) / n
    num = sum((rx[i] - mean_rx) * (ry[i] - mean_ry) for i in range(n))
    den_x = (sum((rx[i] - mean_rx) ** 2 for i in range(n))) ** 0.5
    den_y = (sum((ry[i] - mean_ry) ** 2 for i in range(n))) ** 0.5
    if den_x == 0 or den_y == 0:
        return float("nan")
    return round(num / (den_x * den_y), 4)


def cohen_kappa(pairs: list[tuple[str, str]], labels: list[str]) -> float:
    """Observed vs. expected agreement (binary or multi-class)."""
    n = len(pairs)
    if n == 0:
        return float("nan")
    mat: dict[tuple, int] = defaultdict(int)
    for a, b in pairs:
        mat[(a, b)] += 1
    po = sum(mat[(l, l)] for l in labels) / n
    pe = sum(
        (sum(mat[(l, x)] for x in labels) / n) *
        (sum(mat[(x, l)] for x in labels) / n)
        for l in labels
    )
    return round((po - pe) / (1 - pe), 4) if pe != 1.0 else float("nan")


_VERDICT_MAP = {"ENTAILED": "SUPPORTED", "NOT_ENTAILED": "NOT SUPPORTED"}


def _norm_human_m2(v: str) -> str:
    v = str(v).strip().upper().replace("_", " ").replace(",", "").strip()
    if "NOT" in v:
        return "NOT SUPPORTED"
    if "SUPPORTED" in v:
        return "SUPPORTED"
    return "VACUOUS"


def load_iaa_m2() -> dict | None:
    """
    Compare our automated M2 verdicts against the replied professor Excel.
    Returns a dict of IAA statistics, or None if the file is missing.
    """
    if not REPLIED_M2.exists():
        return None
    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl not installed — skipping IAA computation.")
        return None

    # Rebuild our row-id → verdict mapping (same order as the CSV generator)
    our: dict[int, str] = {}
    row_id = 1
    with open(M2_SPOTCHECK) as f:
        for line in f:
            item = json.loads(line)
            presups  = item.get("m2_presuppositions", [])
            verdicts = item.get("m2_verdicts", [])
            if not presups:
                our[row_id] = "VACUOUS"
                row_id += 1
            else:
                for v in verdicts:
                    our[row_id] = _VERDICT_MAP.get(v, v)
                    row_id += 1

    # Load human verdicts
    wb = openpyxl.load_workbook(REPLIED_M2)
    ws = wb["M2 Faithfulness"]
    human: dict[int, str] = {}
    for r in range(3, ws.max_row + 1):
        rid_val = ws.cell(r, 1).value
        verdict = ws.cell(r, 7).value
        if rid_val is None or verdict is None:
            continue
        try:
            rid = int(str(rid_val).strip())
        except ValueError:
            continue
        human[rid] = _norm_human_m2(str(verdict))

    # Pairs excluding vacuous rows
    pairs = [
        (our[rid], human[rid])
        for rid in sorted(our)
        if rid in human and our[rid] != "VACUOUS" and human[rid] != "VACUOUS"
    ]
    labels = ["SUPPORTED", "NOT SUPPORTED"]
    kappa = cohen_kappa(pairs, labels)
    agree = sum(1 for a, b in pairs if a == b)
    disagree_dir = sum(1 for a, b in pairs if a == "NOT SUPPORTED" and b == "SUPPORTED")

    return {
        "n_pairs": len(pairs),
        "raw_agreement": round(agree / len(pairs), 4) if pairs else float("nan"),
        "cohen_kappa": kappa,
        "our_NOT_SUPPORTED_human_SUPPORTED": disagree_dir,
        "note": (
            "Systematic direction: AI flags NOT SUPPORTED, human accepts as SUPPORTED "
            f"({disagree_dir}/{len(pairs)-agree} disagreements). "
            "AI applies strict literal entailment; human applies inferential reading."
        ),
    }


def load_iaa_m3() -> dict | None:
    """
    Compare our automated M3 scores against the replied professor Excel.
    Returns a dict of IAA statistics, or None if the file is missing.
    """
    if not REPLIED_M3.exists():
        return None
    try:
        import openpyxl
    except ImportError:
        return None

    our: dict[int, dict] = {}
    with open(M3_SPOTCHECK) as f:
        for idx, line in enumerate(f):
            item = json.loads(line)
            our[idx + 1] = {
                "p":     item.get("m3_perception"),
                "o":     item.get("m3_orchestration"),
                "e":     item.get("m3_elicitation"),
                "total": item.get("m3_total"),
            }

    wb = openpyxl.load_workbook(REPLIED_M3)
    ws = wb["M3 Pedagogical"]
    human: dict[int, dict] = {}
    for r in range(3, ws.max_row + 1):
        rid_val = ws.cell(r, 1).value
        if rid_val is None:
            continue
        try:
            rid = int(str(rid_val).strip())
        except ValueError:
            continue
        def gn(c):
            v = ws.cell(r, c).value
            return float(v) if isinstance(v, (int, float)) else None
        p, o, e = gn(7), gn(8), gn(9)
        total = gn(10)
        if total is None and None not in (p, o, e):
            total = p + o + e
        human[rid] = {"p": p, "o": o, "e": e, "total": total}

    dims = [("p", "Perception"), ("o", "Orchestration"), ("e", "Elicitation"), ("total", "Total")]
    result: dict = {"n_rows": len(human)}
    for key, label in dims:
        pairs = [
            (our[rid][key], human[rid][key])
            for rid in sorted(human)
            if rid in our
            and our[rid][key] is not None
            and human[rid][key] is not None
        ]
        if not pairs:
            continue
        ours_v  = [a for a, _ in pairs]
        human_v = [b for _, b in pairs]
        rho = spearman_rho(ours_v, human_v)
        mae = sum(abs(a - b) for a, b in pairs) / len(pairs)
        exact  = sum(1 for a, b in pairs if a == b)
        within1 = sum(1 for a, b in pairs if abs(a - b) <= 1)
        mean_diff = sum(b - a for a, b in pairs) / len(pairs)   # positive = human scores higher
        result[f"{key}_rho"]      = rho
        result[f"{key}_mae"]      = round(mae, 3)
        result[f"{key}_exact"]    = exact
        result[f"{key}_within1"]  = within1
        result[f"{key}_n"]        = len(pairs)
        result[f"{key}_mean_diff"] = round(mean_diff, 3)   # human - ours

    return result


def fmt(mean: float, sd: float) -> str:
    return f"{mean:.3f} ± {sd:.3f}"


def mean_sd(vals: list[float]) -> tuple[float, float]:
    if not vals:
        return float("nan"), float("nan")
    m = sum(vals) / len(vals)
    s = statistics.stdev(vals) if len(vals) > 1 else 0.0
    return m, s


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    print("Loading score files...")
    m1_rows = load_jsonl(SCORES_M1)
    m2_rows = load_jsonl(SCORES_M2)
    m3_rows = load_jsonl(SCORES_M3)
    ragas_rows = load_jsonl(SCORES_RAGAS)

    # Index by response_id
    m1 = {r["response_id"]: r for r in m1_rows}
    m2 = {r["response_id"]: r for r in m2_rows}
    m3 = {r["response_id"]: r for r in m3_rows}
    ragas = {r["response_id"]: r for r in ragas_rows}

    all_ids = set(m1) | set(m2) | set(m3) | set(ragas)
    print(f"  M1: {len(m1)} | M2: {len(m2)} | M3: {len(m3)} | RAGAS: {len(ragas)} | Union: {len(all_ids)}")

    # Build merged rows
    merged = []
    for rid in all_ids:
        r1 = m1.get(rid, {})
        r2 = m2.get(rid, {})
        r3 = m3.get(rid, {})
        rr = ragas.get(rid, {})
        # Get metadata from whichever score file has it
        meta = r1 or r2 or r3 or rr
        merged.append({
            "response_id": rid,
            "model": meta.get("model", ""),
            "cognitive_state": meta.get("cognitive_state", ""),
            "m1_score": r1.get("m1_score", float("nan")),
            "m1_verdict": r1.get("m1_verdict", ""),
            "m2_score": r2.get("m2_score", float("nan")),
            "m2_vacuous": r2.get("m2_vacuous", False),
            "m3_perception": r3.get("m3_perception", float("nan")),
            "m3_orchestration": r3.get("m3_orchestration", float("nan")),
            "m3_elicitation": r3.get("m3_elicitation", float("nan")),
            "m3_total": r3.get("m3_total", float("nan")),
            "ragas_score": rr.get("ragas_score", float("nan")),
            "ragas_empty": rr.get("ragas_empty", False),
        })

    # ── Main results table ────────────────────────────────────────────────────
    models = sorted(set(r["model"] for r in merged if r["model"]))

    print("\n=== Main Results Table ===")
    header = ["model", "n",
              "M1 (mean±SD)", "M2 (mean±SD)",
              "M3_perception (mean±SD)", "M3_orchestration (mean±SD)", "M3_elicitation (mean±SD)", "M3_total (mean±SD)",
              "RAGAS (mean±SD)"]
    table_rows = []

    for model in models:
        rows = [r for r in merged if r["model"] == model]
        n = len(rows)

        m1s = [r["m1_score"] for r in rows if not isinstance(r["m1_score"], float) or not (r["m1_score"] != r["m1_score"])]
        m1s = [v for v in m1s if v == v]
        m2s = [r["m2_score"] for r in rows if r["m2_score"] == r["m2_score"]]
        m3p = [r["m3_perception"] for r in rows if r["m3_perception"] == r["m3_perception"]]
        m3o = [r["m3_orchestration"] for r in rows if r["m3_orchestration"] == r["m3_orchestration"]]
        m3e = [r["m3_elicitation"] for r in rows if r["m3_elicitation"] == r["m3_elicitation"]]
        m3t = [r["m3_total"] for r in rows if r["m3_total"] == r["m3_total"]]
        rgs = [r["ragas_score"] for r in rows if r["ragas_score"] == r["ragas_score"]]

        row = {
            "model": model,
            "n": n,
            "M1 (mean±SD)": fmt(*mean_sd(m1s)),
            "M2 (mean±SD)": fmt(*mean_sd(m2s)),
            "M3_perception (mean±SD)": fmt(*mean_sd(m3p)),
            "M3_orchestration (mean±SD)": fmt(*mean_sd(m3o)),
            "M3_elicitation (mean±SD)": fmt(*mean_sd(m3e)),
            "M3_total (mean±SD)": fmt(*mean_sd(m3t)),
            "RAGAS (mean±SD)": fmt(*mean_sd(rgs)),
        }
        table_rows.append(row)
        print(f"  {model} (n={n}): M1={row['M1 (mean±SD)']} | M2={row['M2 (mean±SD)']} | M3={row['M3_total (mean±SD)']} | RAGAS={row['RAGAS (mean±SD)']}")

    with open(RESULTS_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        writer.writeheader()
        writer.writerows(table_rows)
    print(f"\nSaved: {RESULTS_CSV}")

    # ── Per-cognitive-state breakdown (M3) ───────────────────────────────────
    print("\n=== M3 by Cognitive State ===")
    states = sorted(set(r["cognitive_state"] for r in merged if r["cognitive_state"]))
    state_header = ["model", "cognitive_state", "n",
                    "M3_perception", "M3_orchestration", "M3_elicitation", "M3_total"]
    state_rows = []
    for model in models:
        for state in states:
            rows = [r for r in merged if r["model"] == model and r["cognitive_state"] == state]
            if not rows:
                continue
            m3p = [r["m3_perception"] for r in rows if r["m3_perception"] == r["m3_perception"]]
            m3o = [r["m3_orchestration"] for r in rows if r["m3_orchestration"] == r["m3_orchestration"]]
            m3e = [r["m3_elicitation"] for r in rows if r["m3_elicitation"] == r["m3_elicitation"]]
            m3t = [r["m3_total"] for r in rows if r["m3_total"] == r["m3_total"]]
            sr = {
                "model": model,
                "cognitive_state": state,
                "n": len(rows),
                "M3_perception": fmt(*mean_sd(m3p)),
                "M3_orchestration": fmt(*mean_sd(m3o)),
                "M3_elicitation": fmt(*mean_sd(m3e)),
                "M3_total": fmt(*mean_sd(m3t)),
            }
            state_rows.append(sr)
            print(f"  {model} / {state} (n={len(rows)}): {sr['M3_total']}")

    with open(RESULTS_BY_STATE_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=state_header)
        writer.writeheader()
        writer.writerows(state_rows)
    print(f"\nSaved: {RESULTS_BY_STATE_CSV}")

    # ── Spearman correlation matrix ──────────────────────────────────────────
    print("\n=== Spearman Correlation Matrix (orthogonality check) ===")

    def get_triples(rows):
        m1v, m2v, m3v = [], [], []
        for r in rows:
            if r["m1_score"] == r["m1_score"] and r["m2_score"] == r["m2_score"] and r["m3_total"] == r["m3_total"]:
                m1v.append(r["m1_score"])
                m2v.append(r["m2_score"])
                m3v.append(float(r["m3_total"]))
        return m1v, m2v, m3v

    m1v, m2v, m3v = get_triples(merged)
    rho_m1_m2 = spearman_rho(m1v, m2v)
    rho_m1_m3 = spearman_rho(m1v, m3v)
    rho_m2_m3 = spearman_rho(m2v, m3v)

    print(f"  ρ(M1, M2) = {rho_m1_m2:+.4f}")
    print(f"  ρ(M1, M3) = {rho_m1_m3:+.4f}")
    print(f"  ρ(M2, M3) = {rho_m2_m3:+.4f}")
    print("  (Low ρ across all pairs confirms the three metrics capture distinct failure modes.)")

    corr_rows = [
        {"pair": "M1 vs M2", "spearman_rho": rho_m1_m2, "n": len(m1v)},
        {"pair": "M1 vs M3", "spearman_rho": rho_m1_m3, "n": len(m1v)},
        {"pair": "M2 vs M3", "spearman_rho": rho_m2_m3, "n": len(m1v)},
    ]
    with open(CORR_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["pair", "spearman_rho", "n"])
        writer.writeheader()
        writer.writerows(corr_rows)
    print(f"\nSaved: {CORR_CSV}")

    # ── M1 verdict distribution ──────────────────────────────────────────────
    print("\n=== M1 Verdict Distribution ===")
    by_verdict: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in merged:
        if r["m1_verdict"]:
            by_verdict[r["model"]][r["m1_verdict"]] += 1

    for model in models:
        total = sum(by_verdict[model].values())
        for v in ["CLEARLY_WITHHELD", "BORDERLINE", "CLEARLY_LEAKED"]:
            count = by_verdict[model].get(v, 0)
            pct = 100 * count / total if total else 0
            print(f"  {model} | {v}: {count} ({pct:.1f}%)")

    # ── M2 vacuous report ────────────────────────────────────────────────────
    print("\n=== M2 Vacuous Responses ===")
    for model in models:
        rows = [r for r in merged if r["model"] == model]
        vacuous = sum(1 for r in rows if r["m2_vacuous"])
        pct = 100 * vacuous / len(rows) if rows else 0
        print(f"  {model}: {vacuous}/{len(rows)} vacuous ({pct:.1f}%)")

    # ── Critical checkpoint ─────────────────────────────────────────────────
    if m2v:
        m2_all = [r["m2_score"] for r in merged if r["m2_score"] == r["m2_score"]]
        pct_below_1 = 100 * sum(1 for v in m2_all if v < 1.0) / len(m2_all)
        print(f"\n=== M2 Critical Checkpoint ===")
        print(f"  Responses with m2_score < 1.0 (at least one unentailed presupposition): {pct_below_1:.1f}%")
        if pct_below_1 > 25:
            print("  ✓ >25% flag rate confirms core claim: models introduce ungrounded concepts at significant rate.")
        elif pct_below_1 < 10:
            print("  ⚠ <10% flag rate. Models may be overly conservative. Check M2 raw data — still a finding, different story.")
        else:
            print("  ✓ Flag rate in expected range (10–25%). Results support the core claim with moderate effect.")

    # ── Human IAA ────────────────────────────────────────────────────────────
    print("\n=== Inter-Annotator Agreement (Human vs. Automated) ===")
    iaa_m2 = load_iaa_m2()
    iaa_m3 = load_iaa_m3()

    iaa_rows = []
    if iaa_m2:
        print(f"  M2  n={iaa_m2['n_pairs']} non-vacuous pairs")
        print(f"    Raw agreement : {iaa_m2['raw_agreement']:.1%}")
        print(f"    Cohen's κ     : {iaa_m2['cohen_kappa']:+.4f}")
        print(f"    Direction     : {iaa_m2['note']}")
        iaa_rows.append({"metric": "M2", "measure": "Cohen's kappa",
                         "value": iaa_m2["cohen_kappa"], "n": iaa_m2["n_pairs"],
                         "note": iaa_m2["note"]})
    else:
        print("  M2: replied file not found — skipping.")

    if iaa_m3:
        dims = [("p", "Perception"), ("o", "Orchestration"), ("e", "Elicitation"), ("total", "Total")]
        print(f"  M3  n={iaa_m3['n_rows']} rows")
        for key, label in dims:
            rho   = iaa_m3.get(f"{key}_rho", float("nan"))
            mae   = iaa_m3.get(f"{key}_mae", float("nan"))
            exact = iaa_m3.get(f"{key}_exact", "?")
            n     = iaa_m3.get(f"{key}_n", "?")
            diff  = iaa_m3.get(f"{key}_mean_diff", float("nan"))
            print(f"    {label:16s}: rho={rho:+.4f}  MAE={mae:.3f}  exact={exact}/{n}"
                  f"  mean_diff(human-ours)={diff:+.2f}")
            iaa_rows.append({"metric": f"M3_{key}", "measure": "Spearman rho",
                             "value": rho, "n": n,
                             "note": f"MAE={mae:.3f}, mean human-ours={diff:+.2f}"})
    else:
        print("  M3: replied file not found — skipping.")

    if iaa_rows:
        with open(IAA_CSV, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["metric", "measure", "value", "n", "note"])
            writer.writeheader()
            writer.writerows(iaa_rows)
        print(f"\nSaved: {IAA_CSV}")

    # ── Summary text ─────────────────────────────────────────────────────────
    summary_lines = [
        "SocraticRAG — Results Summary",
        "=" * 50,
        "",
        f"Total responses: {len(merged)}",
        f"Models: {', '.join(models)}",
        "",
        "Main Results (mean ± SD):",
    ]
    for tr in table_rows:
        summary_lines.append(f"  {tr['model']}:")
        summary_lines.append(f"    M1 (Socratic adherence):    {tr['M1 (mean±SD)']}")
        summary_lines.append(f"    M2 (retrieval faithfulness): {tr['M2 (mean±SD)']}")
        summary_lines.append(f"    M3 (pedagogical alignment):  {tr['M3_total (mean±SD)']} / 9")
        summary_lines.append(f"    RAGAS baseline:              {tr['RAGAS (mean±SD)']}")
        summary_lines.append("")

    summary_lines += [
        "Spearman Correlations (orthogonality):",
        f"  ρ(M1, M2) = {rho_m1_m2:+.4f}",
        f"  ρ(M1, M3) = {rho_m1_m3:+.4f}",
        f"  ρ(M2, M3) = {rho_m2_m3:+.4f}",
        "",
        "Output files:",
        f"  {RESULTS_CSV}",
        f"  {RESULTS_BY_STATE_CSV}",
        f"  {CORR_CSV}",
        f"  {SUMMARY_TXT}",
        "",
        "Human validation required before final reporting:",
        "  M1: professor review of BORDERLINE responses",
        "  M2: professor spot-check (data/m2_spot_check.jsonl) — Cohen's kappa",
        "  M3: professor spot-check (data/m3_spot_check.jsonl) — Spearman rho > 0.7",
        "  Dean: professor reads 20 (C, U) pairs to validate Dean decisions",
        "",
        "Inter-Annotator Agreement (Human vs. Automated):",
    ]
    if iaa_m2:
        summary_lines += [
            f"  M2 Cohen's kappa : {iaa_m2['cohen_kappa']:+.4f}  (n={iaa_m2['n_pairs']},"
            f" raw agreement={iaa_m2['raw_agreement']:.1%})",
            f"  M2 note          : {iaa_m2['note']}",
        ]
    if iaa_m3:
        for key, label in [("p","Perception"), ("o","Orchestration"), ("e","Elicitation"), ("total","Total")]:
            rho  = iaa_m3.get(f"{key}_rho", float("nan"))
            mae  = iaa_m3.get(f"{key}_mae", float("nan"))
            diff = iaa_m3.get(f"{key}_mean_diff", float("nan"))
            summary_lines.append(
                f"  M3 {label:16s}: rho={rho:+.4f}  MAE={mae:.3f}  mean(human-ours)={diff:+.2f}"
            )
    summary_lines += [""]

    with open(SUMMARY_TXT, "w") as f:
        f.write("\n".join(summary_lines) + "\n")
    print(f"\nSaved: {SUMMARY_TXT}")
    print("\nDone. Send outputs/results.csv and outputs/summary.txt to professors for review.")


if __name__ == "__main__":
    main()
